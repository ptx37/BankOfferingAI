"""
data_seeder.py — Seeds all 50 customer profiles from the Excel dataset into Redis.

Reads: AI_Hackathon_Product_Offering_Engine_Dataset_v1.xlsx
Stores in Redis:
  - profile:CUST-NNN  → full customer profile + features + events + computed profiling
  - spending:CUST-NNN → top-5 spending categories from transaction history
  - customers:list    → sorted list of all customer IDs with display metadata
"""
from __future__ import annotations

import json
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

EXCEL_PATH = Path(__file__).parent.parent.parent / "AI_Hackathon_Product_Offering_Engine_Dataset_v1.xlsx"

RISK_RANK = {"low": 1, "moderate": 2, "high": 3}

CUSTOMER_NAMES = [
    "Alexandru Ionescu", "Maria Popescu", "Andrei Dumitrescu", "Ioana Georgescu",
    "Bogdan Stanescu", "Elena Constantin", "Mihai Radu", "Diana Matei",
    "Radu Marin", "Cristina Popa", "Daniel Lungu", "Alina Neagu",
    "Florin Moldovan", "Laura Dragomir", "Stefan Ciobanu", "Bianca Roman",
    "Catalin Vasile", "Raluca Oprea", "Marius Dicu", "Andreea Lazar",
    "George Stoica", "Oana Barbu", "Lucian Nica", "Gabriela Tudose",
    "Vlad Petrescu", "Simona Anghel", "Adrian Manolescu", "Corina Florescu",
    "Tudor Cristea", "Mihaela Serban", "Ionut Paun", "Roxana Dima",
    "Cosmin Bucur", "Denisa Voiculescu", "Gabriel Marinescu", "Camelia Nistor",
    "Razvan Badea", "Ana Toma", "Claudiu Iordache", "Monica Enache",
    "Sorin Olaru", "Teodora Zamfir", "Darius Stan", "Mirela Apostol",
    "Cristinel Puscas", "Luminita Voicu", "Silviu Petcu", "Veronica Mocanu",
    "Dorin Codreanu", "Ionela Dobre",
]


def _cid(raw_id: Any) -> str:
    return f"CUST-{int(raw_id):03d}"


def _parse_products(raw: Any) -> list[str]:
    if not raw:
        return []
    parts = [p.strip().lower() for p in str(raw).split(",")]
    return [p for p in parts if p and p != "none"]


def _initials(name: str) -> str:
    parts = name.split()
    return (parts[0][0] + parts[-1][0]).upper() if len(parts) >= 2 else name[:2].upper()


def _segment(income: float) -> str:
    if income >= 7000:
        return "Premium"
    if income >= 3500:
        return "Standard"
    return "Other"


def _financial_health(debt_to_income: float, monthly_savings: float) -> str:
    """PR001-PR003 exact rules."""
    if debt_to_income >= 1.2:
        return "fragile"
    if debt_to_income >= 0.5:
        return "watchlist"
    if monthly_savings > 500:
        return "healthy"
    return "watchlist"


def _investor_readiness(idle_cash: float, monthly_savings: float, events: list[dict]) -> str:
    """PR009-PR011 exact rules."""
    has_trigger = any(e.get("event_type") in ("salary_increase", "bonus") for e in events)
    if idle_cash > 10000 and monthly_savings > 500 and has_trigger:
        return "high"
    if 5000 < idle_cash <= 10000:
        return "medium"
    return "low"


def _context_signals(
    idle_cash: float, monthly_savings: float, investment_gap_flag: int,
    dominant_spend: str, events: list[dict], dependents: int,
    marital_status: str, income: float, homeowner_status: str,
    avg_expenses: float, balance_trend: str, savings: float,
    existing_products: list[str],
) -> list[str]:
    signals: list[str] = []
    event_types = {e.get("event_type", "") for e in events}
    if idle_cash > 10000:
        signals.append("idle_cash_high")
    if "salary_increase" in event_types:
        signals.append("salary_increase")
    if investment_gap_flag == 1:
        signals.append("investment_gap")
    if monthly_savings > 500:
        signals.append("monthly_savings_consistent")
    if "travel_spike" in event_types:
        signals.append("travel_spike")
    if idle_cash > 10000 and dominant_spend not in ("rent",):
        signals.append("inflation_exposed")
    if dominant_spend == "rent":
        signals.append("rent_pattern")
    if dependents > 0 or marital_status.lower() == "married":
        signals.append("family_context")
    if income > 7000:
        signals.append("high_income")
    if dominant_spend == "shopping":
        signals.append("shopping_pattern")
    if "bonus" in event_types:
        signals.append("bonus_event")
    if avg_expenses > 4000:
        signals.append("high_expenses")
    if balance_trend == "declining" and idle_cash < 2000:
        signals.append("liquidity_gap")
    if savings > 10000 and investment_gap_flag == 1:
        signals.append("high_income_no_investments")
    return signals


def _match_score(financial_health: str, investor_readiness: str, signals: list[str]) -> int:
    MUST_HAVE = {"idle_cash_high", "investment_gap", "monthly_savings_consistent",
                 "travel_spike", "rent_pattern", "family_context", "high_income", "high_expenses"}
    base = {"healthy": 65, "watchlist": 45, "fragile": 15}[financial_health]
    r_bonus = {"high": 22, "medium": 12, "low": 3}[investor_readiness]
    s_bonus = min(12, sum(3 for s in signals if s in MUST_HAVE))
    return min(99, base + r_bonus + s_bonus)


def seed_all(redis_client: Any) -> None:
    """Read Excel and seed all customer data into Redis."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return

    if not EXCEL_PATH.exists():
        logger.warning("Excel dataset not found at %s — skipping seed", EXCEL_PATH)
        return

    logger.info("Loading Excel dataset from %s", EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # ── Customers ──────────────────────────────────────────────────────────────
    customers: dict[str, dict] = {}
    ws = wb["customers_enhanced"]
    headers: list[str] | None = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        try:
            raw_id = int(d["customer_id"])
        except (ValueError, TypeError):
            continue
        if raw_id < 1 or raw_id > 50:
            continue
        cid = _cid(raw_id)
        customers[cid] = {
            "customer_id": cid,
            "name": CUSTOMER_NAMES[raw_id - 1],
            "age": int(d.get("age") or 30),
            "city": str(d.get("city") or ""),
            "income": float(d.get("income") or 0),
            "savings": float(d.get("savings") or 0),
            "debt": float(d.get("debt") or 0),
            "has_debt": bool(d.get("has_debt")),
            "risk_profile": str(d.get("risk_profile") or "low").lower(),
            "marital_status": str(d.get("marital_status") or "single").lower(),
            "dependents_count": int(d.get("dependents_count (kids)") or 0),
            "homeowner_status": str(d.get("homeowner_status") or "rent").lower(),
            "existing_products": _parse_products(d.get("existing_products")),
            "profiling_consent": True,
        }

    # ── Features ───────────────────────────────────────────────────────────────
    ws = wb["features_enhanced"]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        try:
            cid = _cid(int(d["customer_id"]))
        except (ValueError, TypeError):
            continue
        if cid not in customers:
            continue
        customers[cid].update({
            "monthly_savings": float(d.get("monthly_savings") or 0),
            "avg_expenses": float(d.get("avg_expenses") or 0),
            "idle_cash": float(d.get("idle_cash") or 0),
            "balance_trend": str(d.get("balance_trend") or "stable"),
            "debt_to_income": float(d.get("debt_to_income") or 0),
            "savings_rate": float(d.get("savings_rate") or 0),
            "dominant_spend_category": str(d.get("dominant_spend_category") or ""),
            "investment_gap_flag": int(d.get("investment_gap_flag") or 0),
        })

    # ── Events ─────────────────────────────────────────────────────────────────
    events_map: dict[str, list[dict]] = defaultdict(list)
    ws = wb["events"]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        try:
            raw_id = int(d.get("customer_id") or 0)
        except (ValueError, TypeError):
            continue
        if raw_id < 1 or raw_id > 50:
            continue
        events_map[_cid(raw_id)].append({
            "event_type": str(d.get("event_type") or ""),
            "date": str(d.get("date") or ""),
        })

    for cid in customers:
        customers[cid]["events"] = events_map.get(cid, [])

    # ── Compute profiling results ───────────────────────────────────────────────
    for cid, c in customers.items():
        fh = _financial_health(c.get("debt_to_income", 0), c.get("monthly_savings", 0))
        ir = _investor_readiness(c.get("idle_cash", 0), c.get("monthly_savings", 0), c.get("events", []))
        rb = c.get("risk_profile", "low")
        sig = _context_signals(
            idle_cash=c.get("idle_cash", 0),
            monthly_savings=c.get("monthly_savings", 0),
            investment_gap_flag=c.get("investment_gap_flag", 0),
            dominant_spend=c.get("dominant_spend_category", ""),
            events=c.get("events", []),
            dependents=c.get("dependents_count", 0),
            marital_status=c.get("marital_status", "single"),
            income=c.get("income", 0),
            homeowner_status=c.get("homeowner_status", "rent"),
            avg_expenses=c.get("avg_expenses", 0),
            balance_trend=c.get("balance_trend", "stable"),
            savings=c.get("savings", 0),
            existing_products=c.get("existing_products", []),
        )
        ms = _match_score(fh, ir, sig)
        c["computed"] = {
            "financial_health": fh,
            "investor_readiness": ir,
            "risk_bucket": rb,
            "context_signals": sig,
            "match_score": ms,
            "segment": _segment(c.get("income", 0)),
            "initials": _initials(c.get("name", cid)),
        }

    # ── Spending from transactions ──────────────────────────────────────────────
    spending_map: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    ws = wb["transactions_enhanced"]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else "" for h in row]
            continue
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        try:
            raw_id = int(d.get("customer_id") or 0)
        except (ValueError, TypeError):
            continue
        if raw_id < 1 or raw_id > 50:
            continue
        cat = str(d.get("category") or "").lower()
        if cat in ("salary", ""):
            continue
        try:
            amt = abs(float(d.get("amount") or 0))
        except (ValueError, TypeError):
            amt = 0.0
        spending_map[_cid(raw_id)][cat] += amt

    wb.close()

    # ── Seed into Redis ─────────────────────────────────────────────────────────
    TTL = 86400 * 30
    customer_list = []
    for cid, profile in sorted(customers.items()):
        redis_client.set(f"profile:{cid}", json.dumps(profile), ex=TTL)

        raw_spending = spending_map.get(cid, {})
        sorted_spending = sorted(raw_spending.items(), key=lambda x: x[1], reverse=True)[:5]
        spending_data = [
            {"category": cat.title(), "amount": round(amt, 2), "isOther": cat == "other"}
            for cat, amt in sorted_spending
        ]
        redis_client.set(f"spending:{cid}", json.dumps(spending_data), ex=TTL)

        comp = profile["computed"]
        customer_list.append({
            "customer_id": cid,
            "name": profile["name"],
            "initials": comp["initials"],
            "segment": comp["segment"],
            "financial_health": comp["financial_health"],
            "risk_profile": profile["risk_profile"],
            "match_score": comp["match_score"],
            "city": profile["city"],
            "age": profile["age"],
            "profiling_consent": profile["profiling_consent"],
        })

    redis_client.set("customers:list", json.dumps(customer_list), ex=TTL)
    logger.info("Seeded %d customer profiles into Redis", len(customers))
